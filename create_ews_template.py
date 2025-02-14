from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

def create_ews_template(filename="EWS_Calculator_Template.xlsx"):
    try:
        wb = Workbook()
        
        # Create sheets
        input_sheet = wb.active
        input_sheet.title = "Input Data"
        results_sheet = wb.create_sheet("Results")
        dashboard_sheet = wb.create_sheet("Dashboard")
        
        # Styles
        header_font = Font(bold=True)
        centered = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Color fills for status
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        orange_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")

        # Create differential styles
        green_dxf = DifferentialStyle(fill=green_fill)
        yellow_dxf = DifferentialStyle(fill=yellow_fill)
        orange_dxf = DifferentialStyle(fill=orange_fill)
        red_dxf = DifferentialStyle(fill=red_fill)
        
        # Input Data Sheet
        input_headers = [
            "Period",
            "Actual Procurement Volume",
            "Target Procurement Volume",
            "Actual Inventory Level",
            "Planned Inventory Level",
            "Outstanding Loan (Other Banks)",
            "Approved Credit Limit",
            "Cash Average Balance",
            "Outstanding Loan Amount",
            "Account Balance",
            "Credit Rating"
        ]
        
        input_sheet['A1'] = "EWS Criteria Calculator"
        input_sheet['A1'].font = Font(bold=True, size=14)
        
        # Set headers
        for col, header in enumerate(input_headers, 1):
            cell = input_sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
            
        # Set column widths
        for col in range(1, len(input_headers) + 1):
            input_sheet.column_dimensions[get_column_letter(col)].width = 20
            
        # Results Sheet
        results_headers = [
            "Period",
            "PVR Status",
            "ILR Status",
            "OLR Status",
            "CLR Status",
            "ABR Status",
            "Credit Status"
        ]
        
        results_sheet['A1'] = "EWS Status Dashboard"
        results_sheet['A1'].font = Font(bold=True, size=14)
        
        # Set headers
        for col, header in enumerate(results_headers, 1):
            cell = results_sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
            
        # Set formulas for row 4
        formulas = [
            "=\'Input Data\'!A4",
            "=IF(ISBLANK('Input Data'!B4), \"\",IF('Input Data'!B4/'Input Data'!C4 >= 0.5, \"Green\",IF('Input Data'!B4/'Input Data'!C4 >= 0.3, \"Yellow\",IF('Input Data'!B4/'Input Data'!C4 >= 0.1, \"Orange\", \"Red\"))))",
            "=IF(ISBLANK('Input Data'!D4), \"\",IF('Input Data'!D4/'Input Data'!E4 <= 1.2, \"Green\",IF('Input Data'!D4/'Input Data'!E4 <= 1.5, \"Yellow\",IF('Input Data'!D4/'Input Data'!E4 <= 1.7, \"Orange\", \"Red\"))))",
            "=IF(ISBLANK('Input Data'!F4), \"\",IF('Input Data'!F4/'Input Data'!G4 <= 0.15, \"Green\",IF('Input Data'!F4/'Input Data'!G4 <= 0.25, \"Yellow\", \"Orange\")))",
            "=IF(ISBLANK('Input Data'!H4), \"\",IF('Input Data'!H4/'Input Data'!I4 >= 0.8, \"Green\",IF('Input Data'!H4/'Input Data'!I4 >= 0.5, \"Yellow\", \"Red\")))",
            "=IF(ISBLANK('Input Data'!J4), \"\",IF('Input Data'!J4/'Input Data'!I4 >= 1, \"Green\", \"Orange\"))",
            "=IF(ISBLANK('Input Data'!K4), \"\",SWITCH('Input Data'!K4,\"Good\", \"Green\",\"Dropped\", \"Yellow\",\"B1\", \"Orange\",\"B2\", \"Orange\",\"B3\", \"Orange\",\"Bad\", \"Red\",\"N/A\"))"
        ]
        
        for col, formula in enumerate(formulas, 1):
            cell = results_sheet.cell(row=4, column=col)
            cell.value = formula
            cell.alignment = centered
            cell.border = border
        
        # Set column widths
        for col in range(1, len(results_headers) + 1):
            results_sheet.column_dimensions[get_column_letter(col)].width = 15
            
        # Add conditional formatting for status colors
        for col in range(2, len(results_headers) + 1):
            col_letter = get_column_letter(col)
            status_range = f"{col_letter}4:{col_letter}100"
            
            # Conditional formatting rules
            rules = [
                ("Green", green_dxf),
                ("Yellow", yellow_dxf),
                ("Orange", orange_dxf),
                ("Red", red_dxf)
            ]
            
            for status, dxf in rules:
                rule = Rule(
                    type="containsText",
                    operator="containsText",
                    text=status,
                    dxf=dxf
                )
                rule.formula = [f'NOT(ISERROR(SEARCH("{status}",{col_letter}4)))']
                results_sheet.conditional_formatting.add(status_range, rule)
        
        # Dashboard Sheet
        dashboard_sheet['A1'] = "EWS Summary Dashboard"
        dashboard_sheet['A1'].font = Font(bold=True, size=14)
        
        dashboard_sheet['A3'] = "Status Summary"
        dashboard_sheet['A3'].font = header_font
        
        status_labels = ["Green", "Yellow", "Orange", "Red"]
        for i, status in enumerate(status_labels, 4):
            dashboard_sheet[f'A{i}'] = status
            dashboard_sheet[f'B{i}'] = f'=COUNTIF(Results!B4:G100, "{status}")'
            
        dashboard_sheet.column_dimensions['A'].width = 15
        dashboard_sheet.column_dimensions['B'].width = 15
        
        # Save the workbook
        wb.save(filename)
        print(f"Excel template has been created: {filename}")
    
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    create_ews_template()